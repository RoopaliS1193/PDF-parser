from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import pdfplumber
from docx import Document
from openpyxl import load_workbook
import pandas as pd
from io import BytesIO, StringIO

app = FastAPI()

# Allow CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Optional: Set max file size (e.g., 5MB)
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

@app.post("/extract-text")
async def extract_text(file: UploadFile = File(...)):
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="PDF too large. Max size is 5MB.")

    try:
        text = ""
        with pdfplumber.open(BytesIO(contents)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return {"text": text.strip()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF parsing failed: {str(e)}")

@app.post("/extract-docx")
async def extract_docx(file: UploadFile = File(...)):
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="DOCX too large. Max size is 5MB.")

    try:
        doc = Document(BytesIO(contents))
        text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        return {"text": text.strip()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DOCX parsing failed: {str(e)}")

@app.post("/extract-xlsx")
async def extract_xlsx(file: UploadFile = File(...)):
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="XLSX too large. Max size is 5MB.")

    try:
        wb = load_workbook(BytesIO(contents), read_only=True)
        text_rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell)
                if row_text.strip():
                    text_rows.append(row_text)
        return {"text": "\n".join(text_rows).strip()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"XLSX parsing failed: {str(e)}")

@app.post("/extract-csv")
async def extract_csv(file: UploadFile = File(...)):
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="CSV too large. Max size is 5MB.")

    try:
        text = contents.decode("utf-8")
        df = pd.read_csv(StringIO(text), dtype=str)
        combined_text = df.fillna("").astype(str).agg(" ".join, axis=1).tolist()
        return {"text": "\n".join(combined_text).strip()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"CSV parsing failed: {str(e)}")
